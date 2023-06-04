import os
import time
import requests
import xlrd
from selene import browser
from selenium import webdriver
from openpyxl import load_workbook
from zipfile import ZipFile
from pypdf import PdfReader

# TODO задание пути к папке resources
download_folder = 'resources'
current_path = os.path.abspath(__file__)
directory_path = os.path.dirname(current_path)
resources_path = os.path.join(directory_path, download_folder)
resources_path = os.path.normpath(resources_path)
print(resources_path)


# TODO работа со скачиванием zip, оформить в тест, добавить ассерты и использовать универсальный путь
def test_download_file_with_browser():
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": resources_path,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)

    browser.config.driver_options = options

    browser.open("https://github.com/Klassik2506/QA_GURU_home_work_lesson5")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(10)

    downloaded_file_path = os.path.join(resources_path, 'QA_GURU_home_work_lesson5-main.zip')
    assert os.path.exists(downloaded_file_path), f"Файл {downloaded_file_path} не скачан!"

    file_size = os.path.getsize(downloaded_file_path)
    assert file_size > 0, f"The file {downloaded_file_path} is empty"


# TODO работа со скачиванием изображений, оформить в тест, сохранять и читать из tmp, использовать универсальный путь
def test_downloaded_file_size():
    url = 'https://w7.pngwing.com/pngs/708/311/png-transparent-icon-logo-twitter-logo-twitter-logo-blue-social-media-area.png'

    r = requests.get(url)
    downloaded_file_path = os.path.join(resources_path, 'Twitter.png')

    with open(downloaded_file_path, 'wb') as file:
        file.write(r.content)

    size = os.path.getsize(downloaded_file_path)

    assert size == 3634


# TODO работа с xlsx, оформить в тест, добавить ассерты и использовать универсальный путь
def test_xlsx():
    xlsx_path = os.path.join(resources_path, 'example_XLSX.xlsx')

    assert os.path.exists(xlsx_path), f'XLSX файл {xlsx_path} не найден'

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active

    assert sheet.cell(row=10, column=8).value == 'Klassik2506'

    print(sheet.cell(row=10, column=8).value)


# TODO работа с xls, оформить в тест, добавить ассерты и использовать универсальный путь
def test_xls():
    xls_file_path = os.path.join(resources_path, 'example_XLS.xls')

    assert os.path.exists(xls_file_path), f'XLS файл {xls_file_path} не найден'

    wb = xlrd.open_workbook(xls_file_path)

    assert wb.nsheets == 2
    assert wb.sheet_names() == ['Лист1', 'Лист2']

    print(f'\nКоличество листов {wb.nsheets}')
    print(f'Имена листов {wb.sheet_names()}')

    sheet = wb.sheet_by_index(0)

    assert sheet.ncols == 3
    assert sheet.nrows == 11
    assert sheet.cell_value(rowx=3, colx=1) == 'Краснов'

    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 3 и столбца 1 = {sheet.cell_value(rowx=3, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))


# TODO Заархивировать имеющиеся в resources различные типы файлов в один архив
def test_zip():
    zip_file_path = os.path.join(resources_path, 'all.zip')

    logo = os.path.join(resources_path, 'Twitter.png')
    pdf_pytest = os.path.join(resources_path, 'PDF_Converter.pdf')
    file_xls = os.path.join(resources_path, 'example_XLS.xls')
    file_xlsx = os.path.join(resources_path, 'example_XLSX.xlsx')


    files = [logo, pdf_pytest, file_xls, file_xlsx]

    with ZipFile(zip_file_path, 'w') as zipF:
        for file in files:
            zipF.write(file, os.path.relpath(file, resources_path))

    with ZipFile(zip_file_path) as zipR:
        for file in files:
            unzipped_path = zipR.namelist()[files.index(file)]
            assert unzipped_path == os.path.relpath(file, resources_path)
            assert os.path.getsize(os.path.join(resources_path, unzipped_path)) == os.path.getsize(file)


# TODO работа с PDF, оформить в тест, добавить ассерты и использовать универсальный путь
def test_pdf():
    pdf_file_path = os.path.join(resources_path, 'PDF_Converter.pdf')

    assert os.path.exists(pdf_file_path), f'PDF файл {pdf_file_path} не найден'

    reader = PdfReader(pdf_file_path)

    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()

    assert number_of_pages == 59
    assert page is not None, 'Page is empty'
    assert text is not None and len(text.strip()) > 0, 'Text is empty'

    print(f'\n1: {number_of_pages}')
    print(f'2: {page}')
    print(f'3: {text}')